"""
Excel Updater for CFP Data
Updates the CWAS Excel file with scraped CFP data.
"""

import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .data_mapper import (
    map_scraped_to_excel,
    CFP_FIELD_MAPPING,
    CITY_INFO_ROWS,
    YEAR_COLUMNS,
    get_field_summary,
    validate_data
)


class ExcelUpdater:
    """Updates Excel file with scraped CFP data."""

    def __init__(self, file_path: str):
        """
        Initialize the Excel updater.

        Args:
            file_path: Path to the Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.changes_log = []

    def open(self):
        """Open the Excel file for editing."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(str(self.file_path), keep_vba=True)
        self.changes_log = []

    def close(self, save: bool = True):
        """Close the Excel file."""
        if self.workbook:
            if save:
                self.workbook.save(str(self.file_path))
            self.workbook.close()
            self.workbook = None

    def update_city_info(self, state: str, city: str):
        """
        Update city and state information in Data Input sheet.

        Args:
            state: State name
            city: City/ULB name
        """
        if not self.workbook:
            raise RuntimeError("Workbook not open. Call open() first.")

        sheet = self.workbook['Data Input']

        # Update state
        state_row = CITY_INFO_ROWS['state']['row']
        state_col = CITY_INFO_ROWS['state']['col']
        old_state = sheet.cell(row=state_row + 1, column=state_col + 1).value
        sheet.cell(row=state_row + 1, column=state_col + 1, value=state)
        self.changes_log.append(f"State: '{old_state}' -> '{state}'")

        # Update city
        city_row = CITY_INFO_ROWS['city']['row']
        city_col = CITY_INFO_ROWS['city']['col']
        old_city = sheet.cell(row=city_row + 1, column=city_col + 1).value
        sheet.cell(row=city_row + 1, column=city_col + 1, value=city)
        self.changes_log.append(f"City: '{old_city}' -> '{city}'")

    def update_cfp_data(self, scraped_data: Dict, year: int = None) -> Dict:
        """
        Update CFP fields in Data Input sheet with scraped data.

        Args:
            scraped_data: Dictionary with scraped financial data
            year: Financial year (if not in scraped_data)

        Returns:
            Dictionary with update summary
        """
        if not self.workbook:
            raise RuntimeError("Workbook not open. Call open() first.")

        sheet = self.workbook['Data Input']

        # Determine year column
        data_year = year or scraped_data.get('year', 2024)
        year_col = YEAR_COLUMNS.get(data_year, 5) + 1  # +1 for openpyxl 1-based indexing

        # Get cell updates from mapper
        updates = map_scraped_to_excel(scraped_data)

        updated_count = 0
        skipped_count = 0

        for row, col, value in updates:
            if value is not None:
                # Rows from mapper are 0-indexed, openpyxl is 1-indexed
                excel_row = row + 1
                old_value = sheet.cell(row=excel_row, column=year_col).value

                # Update the cell
                sheet.cell(row=excel_row, column=year_col, value=value)

                # Log the change
                field_name = self._get_field_name_by_row(row)
                self.changes_log.append(
                    f"Row {excel_row}, Col {get_column_letter(year_col)} ({field_name}): "
                    f"'{old_value}' -> '{value}'"
                )
                updated_count += 1
            else:
                skipped_count += 1

        # Get field summary
        summary = get_field_summary(scraped_data)

        return {
            'year': data_year,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'total_cfp_fields': summary['total_fields'],
            'completion_percentage': summary['completion_percentage'],
            'changes': self.changes_log.copy()
        }

    def _get_field_name_by_row(self, row: int) -> str:
        """Get field name for a given row number."""
        for category, fields in CFP_FIELD_MAPPING.items():
            for field_key, info in fields.items():
                if info['row'] == row:
                    return info['excel_name']
        return f"Unknown (row {row})"

    def clear_cfp_data(self, year: int):
        """
        Clear CFP data for a specific year.

        Args:
            year: Financial year to clear
        """
        if not self.workbook:
            raise RuntimeError("Workbook not open. Call open() first.")

        sheet = self.workbook['Data Input']
        year_col = YEAR_COLUMNS.get(year, 5) + 1

        # Clear all CFP fields for this year
        for category, fields in CFP_FIELD_MAPPING.items():
            for field_key, info in fields.items():
                excel_row = info['row'] + 1
                sheet.cell(row=excel_row, column=year_col, value='ND')
                self.changes_log.append(f"Cleared Row {excel_row}, Col {get_column_letter(year_col)}")

    def create_backup(self) -> str:
        """
        Create a backup of the Excel file.

        Returns:
            Path to backup file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
        backup_path = self.file_path.parent / backup_name

        import shutil
        shutil.copy2(self.file_path, backup_path)

        return str(backup_path)

    def get_changes_log(self) -> List[str]:
        """Get list of changes made."""
        return self.changes_log.copy()


def update_excel_with_cfp_data(
    file_path: str,
    scraped_data: Dict,
    create_backup: bool = True
) -> Dict:
    """
    Convenience function to update Excel file with CFP data.

    Args:
        file_path: Path to Excel file
        scraped_data: Scraped financial data
        create_backup: Whether to create backup before updating

    Returns:
        Dictionary with update summary
    """
    updater = ExcelUpdater(file_path)

    backup_path = None
    if create_backup:
        backup_path = updater.create_backup()

    try:
        updater.open()

        # Update city info if provided
        if scraped_data.get('state') and scraped_data.get('city'):
            updater.update_city_info(
                scraped_data['state'],
                scraped_data['city']
            )

        # Update CFP data
        result = updater.update_cfp_data(scraped_data)
        result['backup_path'] = backup_path

        # Validate data
        issues = validate_data(scraped_data)
        result['validation_issues'] = issues

        updater.close(save=True)

        return result

    except Exception as e:
        updater.close(save=False)
        raise e


if __name__ == "__main__":
    # Test the updater
    print("Testing Excel Updater...")

    # Sample data
    sample_data = {
        'state': 'Tamil Nadu',
        'city': 'Tiruchirappalli City Municipal Corporation',
        'year': 2024,
        'revenue_income': {
            'tax_revenue': 12500.5,
            'assigned_revenue': 3200.0,
        },
        'revenue_expenditure': {
            'establishment_expenses': 8500.0,
        },
        'assets': {
            'cash_bank_balances': 5600.0,
        }
    }

    file_path = "/content/drive/MyDrive/Municipal_Credit_Assessment/data/CWAS - Creditworthiness DIY Tool-Trichy.xlsm"

    if os.path.exists(file_path):
        print(f"\nUpdating: {file_path}")

        result = update_excel_with_cfp_data(file_path, sample_data, create_backup=True)

        print(f"\nResults:")
        print(f"  Year: {result['year']}")
        print(f"  Updated: {result['updated_count']} fields")
        print(f"  Completion: {result['completion_percentage']:.1f}%")
        print(f"  Backup: {result['backup_path']}")

        print(f"\nChanges:")
        for change in result['changes'][:10]:
            print(f"  {change}")
    else:
        print(f"File not found: {file_path}")
